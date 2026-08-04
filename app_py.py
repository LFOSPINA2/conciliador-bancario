import streamlit as st
import pandas as pd
import tabula
from PyPDF2 import PdfReader, PdfWriter
from openpyxl import load_workbook
import tempfile
import os

st.title("Conciliador Bancario – DISTRINORTE")

st.write("Sube el PDF del banco y el AUXILIAR para generar la conciliación.")

pdf_file = st.file_uploader("PDF del banco", type=["pdf"])
aux_file = st.file_uploader("AUXILIAR.xlsx", type=["xlsx"])

if pdf_file and aux_file:
    st.write("Procesando archivos...")

    # Crear archivos temporales
    with tempfile.TemporaryDirectory() as tmpdir:
        pdf_path_original = os.path.join(tmpdir, "BANCOLOMBIA.pdf")
        pdf_path_unlocked = os.path.join(tmpdir, "BANCOLOMBIA_DESBLOQUEADO.pdf")
        ruta_aux = os.path.join(tmpdir, "AUXILIAR.xlsx")
        ruta_aux2 = os.path.join(tmpdir, "AUXILIAR2.xlsx")
        ruta_final = os.path.join(tmpdir, "CONCILIADO.xlsx")
        excel_extracto = os.path.join(tmpdir, "EXTRACTO.xlsx")

        # Guardar archivos subidos
        with open(pdf_path_original, "wb") as f:
            f.write(pdf_file.read())

        with open(ruta_aux, "wb") as f:
            f.write(aux_file.read())

        # ============================================================
        # 🔓 MANEJO DE PDF CON O SIN CONTRASEÑA
        # ============================================================

        reader = PdfReader(pdf_path_original)

        if reader.is_encrypted:
            st.warning("El PDF está protegido con contraseña.")

            password_input = st.text_input("Ingresa la contraseña del PDF", type="password")

            if password_input:
                try:
                    reader.decrypt(password_input)
                    st.success("PDF desbloqueado correctamente.")
                except:
                    st.error("❌ Contraseña incorrecta. Intenta nuevamente.")
                    st.stop()
        else:
            st.info("El PDF no tiene contraseña. Continuando...")

        # Guardar PDF desbloqueado
        writer = PdfWriter()
        for page in reader.pages:
            writer.add_page(page)

        with open(pdf_path_unlocked, "wb") as f:
            writer.write(f)

        # ============================================================
        # 1️⃣ EXTRACCIÓN Y TRATAMIENTO DEL PDF → EXTRACTO.xlsx
        # ============================================================

        tablas = tabula.read_pdf(pdf_path_unlocked, pages='all', multiple_tables=True, stream=True)
        if not tablas:
            tablas = tabula.read_pdf(pdf_path_unlocked, pages='all', multiple_tables=True, lattice=True)

        df_final = pd.concat(tablas, ignore_index=True)

        df_hoja1 = df_final.iloc[:38].reset_index(drop=True)
        df_hoja2 = df_final.iloc[38:].reset_index(drop=True)

        columnas_a_eliminar = [i for i in range(7) if i < len(df_hoja2.columns)]
        df_hoja2.drop(df_hoja2.columns[columnas_a_eliminar], axis=1, inplace=True)

        df_hoja3 = df_hoja1.iloc[5:].reset_index(drop=True)

        fechas = df_hoja3.iloc[:, 0].astype(str).str.split(' ', n=1, expand=True)
        df_hoja3['A'] = fechas[0]
        df_hoja3['B'] = fechas[1]
        df_hoja3.drop(df_hoja3.columns[0], axis=1, inplace=True)

        df_hoja3 = df_hoja3[['A', 'B'] + [col for col in df_hoja3.columns if col not in ['A', 'B']]]

        columnas_a_borrar = []
        if len(df_hoja3.columns) > 2:
            columnas_a_borrar.append(df_hoja3.columns[2])
        if len(df_hoja3.columns) > 6:
            columnas_a_borrar.append(df_hoja3.columns[6])
        df_hoja3.drop(columns=columnas_a_borrar, inplace=True)

        df_hoja3 = df_hoja3.iloc[:, :6]
        df_hoja3.columns = ['FECHA', 'DESCRIPCIÓN', 'SUCURSAL', 'DCTO.', 'VALOR', 'SALDO']

        df_hoja2_copy = df_hoja2.iloc[:, :6].reset_index(drop=True)
        df_hoja2_copy.columns = ['FECHA', 'DESCRIPCIÓN', 'SUCURSAL', 'DCTO.', 'VALOR', 'SALDO']
        df_extracto = pd.concat([df_hoja3, df_hoja2_copy], ignore_index=True)

        df_extracto = df_extracto[df_extracto['DESCRIPCIÓN'] != 'FIN ESTADO DE CUENTA'].reset_index(drop=True)

        def formatear_fecha(valor):
            try:
                partes = str(valor).replace('.', '/').split('/')
                if len(partes) >= 2:
                    return f"{partes[0].zfill(2)}/{partes[1].zfill(2)}/2026"
                return valor
            except:
                return valor

        df_extracto['FECHA'] = df_extracto['FECHA'].apply(formatear_fecha)

        df_extracto['VALOR'] = (
            df_extracto['VALOR'].astype(str)
            .str.replace(',', '', regex=False)
            .str.replace(' ', '', regex=False)
        )

        df_extracto['VALOR'] = pd.to_numeric(df_extracto['VALOR'], errors='coerce').fillna(0)
        df_extracto['SALDOS_POSITIVOS'] = df_extracto['VALOR'].abs()

        df_extracto.to_excel(excel_extracto, index=False)

        # ============================================================
        # 2️⃣ TRATAMIENTO DEL AUXILIAR → AUXILIAR2.xlsx
        # ============================================================

        wb = load_workbook(ruta_aux)
        ws = wb["ExportarAExcel"]

        filas_a_borrar = []
        for fila in range(2, ws.max_row + 1):
            doc_num = ws.cell(row=fila, column=6).value
            if doc_num in (None, "", " "):
                filas_a_borrar.append(fila)

        for fila in reversed(filas_a_borrar):
            ws.delete_rows(fila, 1)

        tabla = []
        for fila in ws.iter_rows(values_only=True):
            if all(c is None for c in fila):
                continue
            tabla.append(list(fila))

        df_aux = pd.DataFrame(tabla)

        encabezado_idx = None
        for i, fila in enumerate(df_aux.values):
            if "Tercero" in str(fila):
                encabezado_idx = i
                break

        df_aux = pd.DataFrame(tabla[encabezado_idx+1:], columns=tabla[encabezado_idx])

        for col in ["Debitos", "Creditos", "Saldo"]:
            df_aux[col] = pd.to_numeric(df_aux[col], errors="coerce").fillna(0)

        df_aux["SALDOS_POSITIVOS"] = df_aux["Debitos"].abs() + df_aux["Creditos"].abs()

        df_aux.to_excel(ruta_aux2, index=False)

        wb2 = load_workbook(ruta_aux2)
        ws2 = wb2.active
        ws2.delete_cols(1)
        wb2.save(ruta_aux2)

        # ============================================================
        # 3️⃣ CRUCE FINAL ENTRE EXTRACTO Y AUXILIAR2 → CONCILIADO.xlsx
        # ============================================================

        df_extracto = pd.read_excel(excel_extracto)
        df_aux = pd.read_excel(ruta_aux2)

        df_extracto["FECHA_AUX"] = None
        df_extracto["DESCRIPCION_AUX"] = None
        df_extracto["DOCNUM_AUX"] = None
        df_extracto["TIPO_COINCIDENCIA"] = None

        df_aux_temp = df_aux.copy()

        conceptos_bancarios = [
            "COBRO IVA PAGOS AUTOMATICOS",
            "CUOTA PLAN CANAL NEGOCIOS",
            "IMPTO GOBIERNO 4X1000",
            "IVA CUOTA PLAN CANAL NEGOCIOS",
            "SERVICIO PAGO A PROVEEDORES",
            "SERVICIO PAGO DE NOMINA"
        ]

        for i, valor in enumerate(df_extracto["SALDOS_POSITIVOS"]):

            descripcion_extracto = str(df_extracto.at[i, "DESCRIPCIÓN"]).strip().upper()

            exacto = df_aux_temp[df_aux_temp["SALDOS_POSITIVOS"] == valor]
            if not exacto.empty:
                fila = exacto.iloc[0]
                df_extracto.at[i, "FECHA_AUX"] = fila["Fecha"]
                df_extracto.at[i, "DESCRIPCION_AUX"] = fila["Nota"]
                df_extracto.at[i, "DOCNUM_AUX"] = fila["Doc Num"]
                df_extracto.at[i, "TIPO_COINCIDENCIA"] = "COINCIDENCIA EXACTA"
                df_aux_temp = df_aux_temp.drop(fila.name)
            else:
                entero = int(valor)
                sin_dec = df_aux_temp[df_aux_temp["SALDOS_POSITIVOS"].astype(int) == entero]
                if not sin_dec.empty:
                    fila = sin_dec.iloc[0]
                    df_extracto.at[i, "FECHA_AUX"] = fila["Fecha"]
                    df_extracto.at[i, "DESCRIPCION_AUX"] = fila["Nota"]
                    df_extracto.at[i, "DOCNUM_AUX"] = fila["Doc Num"]
                    df_extracto.at[i, "TIPO_COINCIDENCIA"] = "SIN DECIMALES"
                    df_aux_temp = df_aux_temp.drop(fila.name)
                else:
                    encontrado = False
                    base = entero

                    for offset in range(1, 101):
                        for ajuste in [offset, -offset]:
                            valor_buscar = base + ajuste
                            rango = df_aux_temp[df_aux_temp["SALDOS_POSITIVOS"] == valor_buscar]

                            if not rango.empty:
                                fila = rango.iloc[0]
                                df_extracto.at[i, "FECHA_AUX"] = fila["Fecha"]
                                df_extracto.at[i, "DESCRIPCION_AUX"] = fila["Nota"]
                                df_extracto.at[i, "DOCNUM_AUX"] = fila["Doc Num"]
                                df_extracto.at[i, "TIPO_COINCIDENCIA"] = f"MAS O MENOS {offset}"
                                df_aux_temp = df_aux_temp.drop(fila.name)
                                encontrado = True
                                break

                        if encontrado:
                            break

                    if not encontrado:
                        df_extracto.at[i, "TIPO_COINCIDENCIA"] = "NO EXISTE"

            for concepto in conceptos_bancarios:
                if concepto in descripcion_extracto:
                    df_extracto.at[i, "DESCRIPCION_AUX"] = "GASTOS BANCARIOS"
                    df_extracto.at[i, "TIPO_COINCIDENCIA"] = "GASTOS BANCARIOS"
                    break

        df_extracto.to_excel(ruta_final, index=False)

        st.success("✔ CRUCE COMPLETO GENERADO → CONCILIADO.xlsx")

        with open(ruta_final, "rb") as f:
            conciliado_bytes = f.read()

        st.download_button(
            label="⬇ Descargar CONCILIADO.xlsx",
            data=conciliado_bytes,
            file_name="CONCILIADO.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
